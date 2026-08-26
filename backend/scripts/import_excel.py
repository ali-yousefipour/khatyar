#!/usr/bin/env python3
"""
وارد کردن داده‌های اکسل سامانهٔ جامع تاکسیرانی به دیتابیس.

استفاده:
  python import_excel.py lines    "/path/گزارش_اطلاعات_خط....xlsx"
  python import_excel.py drivers  "/path/گزارش_اطلاعات_جامع_کاربران....xlsx"
  python import_excel.py oplic    "/path/گزارش_پروانه_های_بهره_برداری....xlsx"
  python import_excel.py taxilic  "/path/گزارش_پروانه_های_تاکسیرانی....xlsx"
  python import_excel.py bills    "/path/گزارش_پرداخت_فیش....xlsx"

نیازمند: pip install openpyxl psycopg2-binary
اتصال از طریق متغیر محیطی DATABASE_URL.
"""
import sys, os, re
from openpyxl import load_workbook
import psycopg2
from psycopg2.extras import execute_batch

DB = os.environ.get("DATABASE_URL", "postgres://postgres:postgres@localhost:5432/taxi")
ZWNJ = "\u200c"

def nid(v):
    s = re.sub(r"\D", "", str(v or ""))
    return s.zfill(10) if s else None

def hmap(ws):
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}

def cell(row, h, name, d=None):
    i = h.get(name)
    return row[i] if i is not None and i < len(row) and row[i] is not None else d

def import_lines(ws, cur):
    h = hmap(ws); rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = cell(r, h, "کد")
        if code is None: continue
        rows.append((str(code), cell(r, h, "مبدا"), cell(r, h, "مقصد"), cell(r, h, "کارگزاری"),
            cell(r, h, "منطقه شهرداری"), cell(r, h, "منطقه تاکسیرانی"), cell(r, h, "تیپ"),
            cell(r, h, "ویژه"), cell(r, h, "گردشی"), cell(r, h, "وضعیت")))
    execute_batch(cur, """
        INSERT INTO lines(code,origin,destination,broker,municipality_zone,taxi_zone,type,is_special,is_circular,status)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (code) DO UPDATE SET origin=EXCLUDED.origin,destination=EXCLUDED.destination,
          broker=EXCLUDED.broker,status=EXCLUDED.status,type=EXCLUDED.type""", rows, page_size=500)
    return len(rows)

def import_drivers(ws, cur):
    h = hmap(ws); rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = nid(cell(r, h, "کد ملی"))
        if not code: continue
        rows.append((code, cell(r, h, "نام"), cell(r, h, "نام خانوادگی"),
            cell(r, h, "شماره تلفن همراه") or cell(r, h, "تلفن همراه"), cell(r, h, "شماره هوشمند")))
    execute_batch(cur, """
        INSERT INTO drivers(national_id,first_name,last_name,mobile,smart_no,updated_at)
        VALUES (%s,%s,%s,%s,%s,now())
        ON CONFLICT (national_id) DO UPDATE SET first_name=EXCLUDED.first_name,
          last_name=EXCLUDED.last_name, mobile=COALESCE(EXCLUDED.mobile,drivers.mobile),
          smart_no=COALESCE(EXCLUDED.smart_no,drivers.smart_no), updated_at=now()""", rows, page_size=1000)
    return len(rows)

def import_oplic(ws, cur):
    h = hmap(ws); drv = []; veh = []; link = []
    op_code_key = "کد بهره" + ZWNJ + "برداری"
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = nid(cell(r, h, "کد ملی"))
        plate = cell(r, h, "پلاک خودرو")
        opcode = cell(r, h, op_code_key) or cell(r, h, "کد بهره‌برداری")
        if code:
            drv.append((code, cell(r, h, "نام"), cell(r, h, "نام خانوادگی"),
                str(opcode) if opcode else None, cell(r, h, "تاریخ آغاز"),
                cell(r, h, "تاریخ انقضا"), cell(r, h, "وضعیت"), cell(r, h, "نوع راننده")))
        if plate:
            veh.append((str(plate), cell(r, h, "شماره وین خودرو"), cell(r, h, "نوع خودرو"),
                cell(r, h, "مدل خودرو"), str(opcode) if opcode else None, code))
            if code: link.append((str(plate), code))
    execute_batch(cur, """
        INSERT INTO drivers(national_id,first_name,last_name,operating_code,op_lic_issue,op_lic_expire,op_lic_status,driver_type,updated_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,now())
        ON CONFLICT (national_id) DO UPDATE SET operating_code=EXCLUDED.operating_code,
          op_lic_issue=EXCLUDED.op_lic_issue, op_lic_expire=EXCLUDED.op_lic_expire,
          op_lic_status=EXCLUDED.op_lic_status, updated_at=now()""", drv, page_size=1000)
    execute_batch(cur, """
        INSERT INTO vehicles(plate,vin,model_name,model_year,operating_code,owner_national_id)
        VALUES (%s,%s,%s,%s,%s,%s)
        ON CONFLICT (plate) DO UPDATE SET vin=EXCLUDED.vin,model_name=EXCLUDED.model_name,
          operating_code=EXCLUDED.operating_code,owner_national_id=EXCLUDED.owner_national_id""", veh, page_size=1000)
    for plate, code in link:
        cur.execute("""INSERT INTO vehicle_drivers(vehicle_id,driver_id,role)
            SELECT v.id,d.id,'beneficiary' FROM vehicles v, drivers d
            WHERE v.plate=%s AND d.national_id=%s
            ON CONFLICT (vehicle_id,driver_id,shift) DO NOTHING""", (plate, code))
    return len(drv)

def import_taxilic(ws, cur):
    h = hmap(ws); rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = nid(cell(r, h, "کد ملی"))
        if not code: continue
        rows.append((code, cell(r, h, "نام"), cell(r, h, "نام خانوادگی"), cell(r, h, "شماره هوشمند"),
            cell(r, h, "تاریخ شروع"), cell(r, h, "تاریخ انقضا"), cell(r, h, "وضعیت")))
    execute_batch(cur, """
        INSERT INTO drivers(national_id,first_name,last_name,smart_no,taxi_lic_issue,taxi_lic_expire,taxi_lic_status,updated_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,now())
        ON CONFLICT (national_id) DO UPDATE SET smart_no=COALESCE(EXCLUDED.smart_no,drivers.smart_no),
          taxi_lic_issue=EXCLUDED.taxi_lic_issue, taxi_lic_expire=EXCLUDED.taxi_lic_expire,
          taxi_lic_status=EXCLUDED.taxi_lic_status, updated_at=now()""", rows, page_size=1000)
    return len(rows)

def import_bills(ws, cur):
    h = hmap(ws); batch = []; total = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        batch.append((cell(r, h, "شناسه قبض"), cell(r, h, "شناسه پرداخت"), cell(r, h, "وضعیت پرداخت") or cell(r, h, "وضعیت"),
            cell(r, h, "بابت"), cell(r, h, "عنوان شخص/شرکت"), nid(cell(r, h, "کد/شناسه ملی") or cell(r, h, "کد ملی")),
            cell(r, h, "تلفن شخص/شرکت"), cell(r, h, "مبلغ"), cell(r, h, "تاریخ پرداخت"), cell(r, h, "پلاک")))
        if len(batch) >= 2000: total += _flush_bills(cur, batch); batch = []
    total += _flush_bills(cur, batch); return total

def _flush_bills(cur, batch):
    if not batch: return 0
    execute_batch(cur, """
        INSERT INTO bills(bill_id,pay_id,status,reason,person_title,national_id,phone,amount,pay_date,plate)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", batch, page_size=2000)
    return len(batch)

IMPORTERS = {"lines": import_lines, "drivers": import_drivers, "oplic": import_oplic,
             "taxilic": import_taxilic, "bills": import_bills}

def main():
    if len(sys.argv) != 3 or sys.argv[1] not in IMPORTERS:
        print(__doc__); sys.exit(1)
    kind, path = sys.argv[1], sys.argv[2]
    wb = load_workbook(path, read_only=True); ws = wb.active
    conn = psycopg2.connect(DB); cur = conn.cursor()
    n = IMPORTERS[kind](ws, cur)
    conn.commit(); cur.close(); conn.close(); wb.close()
    print(f"{n} رکورد {kind} وارد شد.")

if __name__ == "__main__":
    main()
